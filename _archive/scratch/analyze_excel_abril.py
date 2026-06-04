"""Analisa planilha Excel manual vs siglas do sistema."""
import sys
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

import openpyxl
import pandas as pd

EXCEL = Path(r"C:\Users\xirin\OneDrive\Desktop\Escala Piloto de Apoio\PAO APAO - ABRIL 2026.xlsx")
SHEET = "Abril - Publicada"

# Siglas Excel -> tipo interno (sistema mantém FER/C/FP/V/S na UI)
EXCEL_TO_SYSTEM = {
    "FR": "FOLGA",
    "F": "FOLGA",
    "FP": "FOLGA PEDIDA",
    "FS": "FOLGA SOCIAL",
    "FA": "FOLGA AGRUPADA",  # confirmar com usuário
    "FB": "? FOLGA (FB na legenda)",
    "FF": "? FOLGA (FF na legenda)",
    "FC": "FOLGA COMPENSA (legenda)",
    "L": "FÉRIAS → sistema: FER",
    "K": "CURSO → sistema: C",
    "V": "VOO",
    "SM": "SIMULADOR → sistema: S",
    "S": "SIMULADOR",
    "ND": "ND",
    "DM": "DISPENSA MÉDICA",
    "EP": "? EXAME PERIÓDICO (produtivo na planilha?)",
    "AV": "? (1 ocorrência Lucaas)",
    "6": "T6",
    "7": "T7",
    "8": "T8",
    "9": "T9",
    "0": "APAO T4?",
    "1": "APAO T1",
    "2": "APAO T2",
    "3": "APAO T3",
    "4": "APAO T4",
}

wb = openpyxl.load_workbook(EXCEL, data_only=False)
ws = wb[SHEET]

print("=== LEGENDA ABRIL (colunas A-B e E-G) ===")
for r in range(31, 46):
    parts = []
    for c in (1, 2, 5, 6, 7):
        v = ws.cell(r, c).value
        if v is not None:
            parts.append(f"c{c}={v!r}")
    if parts:
        print(f"row {r}: ", " | ".join(parts))

print("\n=== COUNTIF patterns (resumo cobertura) ===")
ctr = Counter()
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, str) and "COUNTIF" in v.upper():
            suffix = v.split(",")[-1].rstrip(")")
            key = f"COUNTIF(...,{suffix}"
            ctr[key] += 1
for k, n in ctr.most_common(20):
    print(f"  {n:3d}x  {k}")

# rows 22-27 labels in col B?
print("\n=== Linhas de contagem (col A/B) ===")
for r in range(22, 32):
    a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
    f = ws.cell(r, 6).value
    if a or b or (isinstance(f, str) and "COUNTIF" in f):
        print(f"  r{r}: A={a!r} B={b!r} F={f!r}")

# Regras inferidas dos pilotos PAO (meta ~20 prod / 10 folgas)
df = pd.read_excel(EXCEL, sheet_name=SHEET, header=None)
PRODUCTIVE = {"6", "7", "8", "9", "V", "K", "SM", "S", "EP"}
REST = {"FR", "FP", "FS", "FA", "FB", "FF", "FC"}

def pilot_stats():
    hdr = df.iloc[2]
    day_cols = []
    for j, v in enumerate(hdr):
        try:
            d = int(float(v))
            if 1 <= d <= 31 or d in (30, 31):
                day_cols.append((j, d))
        except Exception:
            pass
    seen = set()
    dc = []
    for j, d in sorted(day_cols, key=lambda x: x[0]):
        if j not in seen:
            seen.add(j)
            dc.append((j, d))

    print("\n=== METAS PLANILHA (PAO regular, Abril Publicada) ===")
    for i in range(5, len(df)):
        name = str(df.iloc[i, 0]).strip()
        if not name or name.lower() == "nan":
            continue
        if any(x in name.upper() for x in ("QUANTIDADE", "TURNO", "HOR", "LEGEND", "CMD")):
            continue
        if "FCF" in name.upper():
            continue
        cells = {}
        for j, d in dc:
            v = df.iloc[i, j]
            if pd.isna(v):
                continue
            c = str(v).strip().upper().replace(".0", "")
            if c not in ("NAN", ""):
                cells[d] = c
        if not any(c in {"6", "7", "8", "V", "K", "SM"} for c in cells.values()):
            if "ANTONIO" not in name.upper():
                continue
        prod = sum(1 for c in cells.values() if c in PRODUCTIVE)
        rest = sum(1 for c in cells.values() if c in REST)
        ferias = sum(1 for c in cells.values() if c == "L")
        print(f"  {name[:28]:28} prod={prod:2d} folgas={rest:2d} ferias={ferias:2d}")

pilot_stats()

print("\n=== DÚVIDAS DE SIGLA (confirmar com usuário) ===")
for code, meaning in sorted(EXCEL_TO_SYSTEM.items()):
    if "?" in meaning or code in ("FA", "FB", "FF", "FC", "EP", "AV"):
        print(f"  Excel {code:4} → {meaning}")
